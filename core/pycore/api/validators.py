# | school project | Jakub Lemiesiewicz |
# | Zespół Szkół Komunikacji w Poznaniu |
#  from django.core.exceptions import ValidationError

from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from .exceptions import Error, Errors


def validate_xlsx_header(request_data, header):  # noqa
    pass


def validate_xlsx_file(file_obj):
    XLSX_ERROR_CATEGORY = 'xlsx cell error value'  # TODO WSP-84 create list of errors
    workbook = load_workbook(file_obj)
    worksheet = workbook.active
    size = worksheet.max_row + 1

    errors = list()
    correct = True

    if worksheet[f'A{1}'].value != 'max_weight':
        correct = False
        errors.append(Error(category=XLSX_ERROR_CATEGORY, message="header error A1"))

    if worksheet[f'B{1}'].value != 'price':
        correct = False
        errors.append(Error(category=XLSX_ERROR_CATEGORY, message="header error B1"))

    for i in range(2, size, 1):
        try:
            float(worksheet[f'A{i}'].value)
        except Exception as e:  # noqa
            errors.append(Error(category=XLSX_ERROR_CATEGORY, message=f'row error row {i} col A'))
            correct = False

        try:
            float(worksheet[f'B{i}'].value)
        except Exception as e:  # noqa
            errors.append(Error(category=XLSX_ERROR_CATEGORY, message=f'row error row {i} col B'))
            correct = False

    if not correct:
        detail = Errors(errors)
        raise ValidationError(detail.json_array)

def validate_load_start_data(data):  # noqa:W0613
    # raise ValidationError({"test": "test_test"})
    return data


def validate_product(data):
    # TODO WSP-65
    return data


def validate_producer(data):
    # TODO WSP-65
    return data


def validate_category(data):
    # TODO WSP-65
    return data
