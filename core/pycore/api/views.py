# | school project | Jakub Lemiesiewicz |
# | Zespół Szkół Komunikacji w Poznaniu |
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import action

from openpyxl import load_workbook
from django.shortcuts import get_object_or_404
from .models import Note, PriceList, Option, Producer, Product, Category
from . import serializers, validators, exceptions


class GetTestViewSet(APIView):
    def get(self, request):  # noqa:W0613
        return Response("OK")


class LoadStartDataViewSet(APIView):
    MAIN_PRICELIST_IDENTIFIER = 'mainwsppricelist'

    def check(self):
        try:
            main_pricelist = PriceList.objects.get(main_identifier=self.MAIN_PRICELIST_IDENTIFIER)  # noqa:W0612
            return False
        except Exception as e:  # noqa:W0612
            return True
        return True

    def post(self, request):
        validators.validate_load_start_data(request)

        # TODO WSP-42 TOKEN VALIDATE
        code = "The pricelist data has already existed on the instance"
        if self.check():

            pricelistarray = [
                {'max_weight': 10, 'price': 9},
                {'max_weight': 35, 'price': 19},
                {'max_weight': 50, 'price': 29},
                {'max_weight': 65, 'price': 49},
                {'max_weight': 80, 'price': 59},
                {'max_weight': 90, 'price': 69},
                {'max_weight': 100, 'price': 79},
                {'max_weight': 125, 'price': 99},
                {'max_weight': 150, 'price': 119},
                {'max_weight': 175, 'price': 199},
                {'max_weight': 20000000, 'price': 249},
            ]

            pricelist = PriceList.objects.create(
                main_identifier=self.MAIN_PRICELIST_IDENTIFIER,
                quantity=len(pricelistarray)
            )
            pricelist.save()

            for option in pricelistarray:
                o = Option.objects.create(max_weight=option['max_weight'], price=option['price'], price_list=pricelist)
                o.save()
                code = "The pricelist data has been loaded correctly"

        return Response({"status": "OK", "code": code})


class NoteViewSet(ModelViewSet):  # noqa:R0901
    serializer_class = serializers.NoteSerializer
    queryset = Note.objects.all()


class PriceListViewSet(ModelViewSet):  # noqa:R0901
    serializer_class = serializers.PriceListSerializer
    queryset = PriceList.objects.all()
    lookup_field = 'main_identifier'
    lookup_url_kwarg = 'main_identifier'

    def retrieve(self, request, main_identifier=None):  # noqa:W0221
        queryset = PriceList.objects.all()
        pricelist = get_object_or_404(queryset, main_identifier=main_identifier)
        serializer = serializers.PriceListSerializer(pricelist)

        response = {
            'price_list': serializer.data['main_identifier'],
            'options': list()
        }

        for pk in serializer.data['options']:
            option = get_object_or_404(Option, pk=pk)
            response['options'].append(
                option.json_view
            )

        return Response(response)

    @action(detail=False, methods=['post'], serializer_class=serializers.XlsxPriceListSerializer)
    def upload_mainwsppricelist(self, request):
        serializer = serializers.XlsxPriceListSerializer(data=request.data)
        if not serializer.is_valid():
            raise exceptions.InvalidRequest(serializer.errors)

        def delete():
            try:
                mainwsppricelist = get_object_or_404(PriceList, main_identifier='mainwsppricelist')
                mainwsppricelist.delete()
            except Exception as e:  # noqa
                pass

        delete()

        mainwsppricelist = PriceList(main_identifier='mainwsppricelist', quantity=0)
        mainwsppricelist.save()

        file_object = serializer.validated_data['pricelistfile']
        workbook = load_workbook(file_object)
        worksheet = workbook.active
        worksheet_size = worksheet.max_row + 1
        options = list()

        for i in range(2, worksheet_size, 1):
            option_max_weight = worksheet[f'A{i}'].value
            option_price = worksheet[f'B{i}'].value
            option = Option(max_weight=option_max_weight, price=option_price, price_list=mainwsppricelist)
            option.save()
            options.append(option)

        mainwsppricelist.quantity = len(options)
        mainwsppricelist.save()
        return Response({"OK": len(options)})


class OptionViewSet(ModelViewSet):  # noqa:R0901
    serializer_class = serializers.OptionSerializer
    queryset = Option.objects.all()


class ProductViewSet(ModelViewSet):
    serializer_class = serializers.ProductSerializer
    queryset = Product.objects.all()


class ProducerViewSet(ModelViewSet):
    serializer_class = serializers.ProducerSerializer
    queryset = Producer.objects.all()


class CategoryViewSet(ModelViewSet):
    serializer_class = serializers.CategorySerializer
    queryset = Category.objects.all()
